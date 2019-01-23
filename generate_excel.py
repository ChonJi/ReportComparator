import comparator_runner
from report_provider import ReportProvider
import openpyxl

def getNumbersFromName(name):
    s = []
    for c in name:
        try:
            int(c, 10)
            s.append(c)
        except ValueError:
            continue
    return ''.join(c for c in s)

if __name__ == "__main__":
    #ReportProvider()
    c = comparator_runner.ComparatorRunner(['SKTB-55194', 'SKTB-31396 Chorus SW Flashing'])
    wb = openpyxl.Workbook()
    sheet = wb.active
    #sheet.cell(row=2, column=3)
    sheet.cell(row=1, column=1).value = 'Test  |  Pipeline'
    for index, log in enumerate(c.logs_paths):
        sheet.cell(row=1, column=index+2).value = int(getNumbersFromName(log))
    row_id = 2
    if c.important is not None:
        for imp in c.important:
            if imp in c.filtered_overview:
                sheet.cell(row=row_id, column=1).value = imp
                for idx, result in enumerate(c.filtered_overview[imp]):
                    status = result
                    if status == 'PASS':
                        status = 'Passed'
                    elif status == 'FAIL':
                        status = 'Failed'
                    sheet.cell(row=row_id, column=idx + 2).value = status
                row_id+=1

    for test in c.filtered_overview:
        if test in c.important:
            continue
        sheet.cell(row=row_id, column=1).value = test
        for idx, result in enumerate(c.filtered_overview[test]):
            status = result
            if status=='PASS':
                status='Passed'
            elif status=='FAIL':
                status='Failed'
            sheet.cell(row=row_id, column=idx+2).value = status
        row_id+=1


    wb.save('temp.xlsx')
